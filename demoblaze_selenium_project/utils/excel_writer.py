# utils/excel_writer.py
import os
from openpyxl import Workbook
from datetime import datetime

def write_to_excel(test_name, status, message):
    folder = "xl_reports"
    os.makedirs(folder, exist_ok=True)

    file_path = os.path.join(folder, "test_results.xlsx")
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        ws.append(["Timestamp", "Test Name", "Status", "Message"])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, test_name, status, message])
    wb.save(file_path)
