import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


# ✅ Create screenshots folder and save image with timestamp
def take_screenshot(driver, name):
    os.makedirs("screenshots", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"screenshots/{name}_{timestamp}.png"
    driver.save_screenshot(filename)
    return filename  # You can optionally log/print this in your test


# ✅ Create or update Excel report in xl_report/test_results.xlsx
def log_to_excel(test_name, status, log_message):
    os.makedirs("xl_report", exist_ok=True)
    file_path = "xl_report/test_results.xlsx"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 📄 Create workbook if not exists
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        ws.append(["Test Name", "Status", "Log Message", "Timestamp"])
        # Apply bold font to headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(file_path)

    # ✏️ Append new test result
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([test_name, status, log_message, timestamp])
    wb.save(file_path)

    print(f"✅ Logged to Excel: {test_name} - {status}")

