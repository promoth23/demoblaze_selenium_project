import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Ensure xl_reports/ exists
os.makedirs("xl_reports", exist_ok=True)

def log_test_result(test_name, status, message):
    file_path = "xl_reports/test_results.xlsx"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(file_path):
        # Create a new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Test Name", "Status", "Message"])
        wb.save(file_path)

    # Append test result
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([timestamp, test_name, status, message])
    wb.save(file_path)
